"""
Fig_S1c_hey1_sensitivity_analysis.py
=====================================
Hey1-for-Hes1 normalization sensitivity reanalysis (Supplementary Figure 1c /
Supplementary Table 2).

Tests whether the 27-pair consensus co-expression network identified under the
original normalization scheme (raw + /Hes1 + /Kcnj11 + /Ins1) is robust to
substitution of Hey1 for Hes1 as the Notch-target normalization denominator.

Method
------
A gene pair (g1, g2) is analytically validated if |r| >= 0.70 with consistent
sign across every normalization condition in which NEITHER gene serves as the
denominator.  This matches the recurrence criterion used for the 27-pair
consensus network reported in the manuscript.

The script evaluates two normalization schemes:
  Original  : raw + /Hes1 + /Kcnj11 + /Ins1
  Alternative: raw + /Hey1 + /Kcnj11 + /Ins1

and reports overlap, pairs unique to each scheme, and the retention rate.

Input
-----
  Supplementary_Data_1.xlsx — sheet "Supplementary Data 1"
  Rows 5+ : gene name (col A) + 12 replicate dCt values
  Columns  : CTRL (cols G–J), DAPT (cols M–P), DKK-1 (cols S–V)
  Set DATA_FILE below to the path of this file.

Dependencies: openpyxl, numpy, scipy
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
import numpy as np
from scipy.stats import pearsonr
from itertools import combinations

# ── User-configurable path ────────────────────────────────────────────────────
DATA_FILE = "Supplementary_Data_1.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(DATA_FILE)
ws = wb["Supplementary Data 1"]

raw_data = {}
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    gene = row[0]
    if not gene or not isinstance(gene, str):
        continue
    if gene.startswith(("Abbreviations", "Trend", "CTRL", "Hey1,", "†", "Ins2")):
        continue
    vals = row[6:10] + row[12:16] + row[18:22]   # 4 CTRL + 4 DAPT + 4 DKK-1 replicates
    if any(v is None or not isinstance(v, (int, float)) for v in vals):
        continue
    raw_data[gene] = np.array(vals, dtype=float)

all_genes = sorted(raw_data.keys())
print(f"Genes loaded ({len(all_genes)}): {all_genes}\n")

# ── Validated-pair computation ────────────────────────────────────────────────
THRESHOLD = 0.70

def compute_validated_pairs(norm_gene_list):
    """
    Return a dict of validated pairs under the given normalization scheme.

    norm_gene_list: list of denominator genes (e.g. ['Hes1', 'Kcnj11', 'Ins1'])
    Conditions tested: 'raw' plus one per denominator gene (denominator gene
    excluded from pairwise evaluation in its own condition).
    """
    conditions = {"raw": {g: raw_data[g].copy() for g in all_genes}}
    for ng in norm_gene_list:
        denom = raw_data[ng]
        conditions[ng] = {g: raw_data[g] / denom for g in all_genes}

    validated = {}
    for g1, g2 in combinations(all_genes, 2):
        applicable = []
        for cname, cdata in conditions.items():
            if cname == "raw" or (cname != g1 and cname != g2):
                applicable.append((cname, cdata))

        if not applicable:
            continue

        rs = []
        valid = True
        for cname, cdata in applicable:
            x, y = cdata[g1], cdata[g2]
            if np.std(x) < 1e-10 or np.std(y) < 1e-10:
                valid = False; break
            r, _ = pearsonr(x, y)
            if np.isnan(r) or abs(r) < THRESHOLD:
                valid = False; break
            rs.append(r)

        if valid and rs and (all(r > 0 for r in rs) or all(r < 0 for r in rs)):
            validated[frozenset([g1, g2])] = {
                "rs": rs,
                "applicable_conds": [c[0] for c in applicable],
                "mean_r": float(np.mean(rs)),
            }
    return validated

# ── Original scheme ───────────────────────────────────────────────────────────
print("=== ORIGINAL: raw + /Hes1 + /Kcnj11 + /Ins1 ===")
orig = compute_validated_pairs(["Hes1", "Kcnj11", "Ins1"])
print(f"Validated pairs: {len(orig)}")
for p, v in sorted(orig.items(), key=lambda x: -abs(x[1]["mean_r"])):
    gl = sorted(p)
    rs_str = ", ".join(f"{r:.3f}" for r in v["rs"])
    print(f"  {gl[0]:12s} — {gl[1]:12s}  [{', '.join(v['applicable_conds'])}]"
          f"  r=[{rs_str}]  mean={v['mean_r']:.3f}")

# ── Alternative scheme (Hey1 replaces Hes1) ───────────────────────────────────
print("\n=== ALTERNATIVE: raw + /Hey1 + /Kcnj11 + /Ins1 ===")
alt = compute_validated_pairs(["Hey1", "Kcnj11", "Ins1"])
print(f"Validated pairs: {len(alt)}")
for p, v in sorted(alt.items(), key=lambda x: -abs(x[1]["mean_r"])):
    gl = sorted(p)
    rs_str = ", ".join(f"{r:.3f}" for r in v["rs"])
    print(f"  {gl[0]:12s} — {gl[1]:12s}  [{', '.join(v['applicable_conds'])}]"
          f"  r=[{rs_str}]  mean={v['mean_r']:.3f}")

# ── Overlap summary ───────────────────────────────────────────────────────────
o = set(orig.keys()); a = set(alt.keys())
shared  = o & a
only_o  = o - a
only_a  = a - o

print(f"\n=== OVERLAP SUMMARY ===")
print(f"Original pairs:             {len(o)}")
print(f"Alternative (Hey1) pairs:   {len(a)}")
print(f"Shared (retained):          {len(shared)}  ({len(shared)/len(o)*100:.1f}% of original)")
print(f"Lost (original only):       {len(only_o)}")
print(f"Gained (alternative only):  {len(only_a)}")

if only_o:
    print("\n  Pairs lost when Hes1 is replaced by Hey1:")
    for p in sorted(only_o):
        gl = sorted(p); print(f"    {gl[0]} — {gl[1]}")

if only_a:
    print("\n  New pairs appearing only under Hey1 normalization:")
    for p in sorted(only_a):
        gl = sorted(p); print(f"    {gl[0]} — {gl[1]}")

# ── Ins2 availability note ────────────────────────────────────────────────────
print(f"\nIns2 measured in RT-qPCR panel: {'Ins2' in raw_data}")
print("(Ins2 substitution for Ins1 would require new RT-qPCR experiments.)")
